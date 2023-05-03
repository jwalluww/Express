#**********************************#
#***** Walmart US Update SKUs *****#
#***** Listed - wm-us-listed  *****#
#**********************************#
#-- Walmart US Price & Inventory
#-- Automated Hourly
#-- Prints to Listed & Listed Catalog

def wmus_listed():
    
    
    
    ########################
    ### Import Libraries ###
    ########################
    #%%
    # Base Imports
    import pandas as pd
    import numpy as np
    import os
    from datetime import datetime
    import time
    import pytz
    from openpyxl import load_workbook
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    from googleapiclient.http import MediaFileUpload
    
    # Primary folder for scripts & passcodes
    folder = "G:\\.shortcut-targets-by-id\\1-T64nMCJ6WrtE7nvU7MzJPa6nzDxRw8c\\Justin&Carmen\\Carmen data\\Automation"
    
    # Change the current directory to specified directory
    os.chdir(folder)
    
    # Import Business Functions
    import biz_funcs as bf
    
    # Set Directories & Passcodes to Google
    folder,drive,client,bqclient,credentials,project_id = bf.google_credentials()
    #%%
    
    
    
    
    
    #############################################
    ### Import and join Manual Entry Database ###
    #############################################
    #%%
    # Read
    dbmn = bf.read_gsheet(client=client, worksheet="Listed Catalog", tab='values', twoheader=True)
    
    # Rename/Format
    dbmn['couponRate']              = dbmn['Coupon Rate'].replace('',0).str.rstrip("%").astype(float)/100
    dbmn['referralRate']            = dbmn['Referral Rate'].replace('',0).str.rstrip("%").astype(float)/100
    dbmn['buyBoxRepricePercent']    = dbmn['Buy Box Reprice Percent'].replace('',0).str.rstrip("%").astype(float)/100
    dbmn['minMarkup']               = dbmn['Min Markup'].replace('',0).str.rstrip("%").astype(float)/100
    dbmn['maxMarkup']               = dbmn['Max Markup'].replace('',0).str.rstrip("%").astype(float)/100
    dbmn['discountRateforGiftCard'] = dbmn['Discount Rate for Gift Card'].replace('',0).str.rstrip("%").astype(float)/100
    dbmn['taxPercent']              = dbmn['Tax percent'].replace('',0).str.rstrip("%").astype(float)/100
    dbmn['otherCosts']              = dbmn['Other Costs'].replace('',0).str.strip("$").astype(float)
    dbmn['subscriptionDiscount']    = dbmn['Subscription discount'].replace('',0).str.rstrip("%").astype(float)/100
    dbmn = dbmn.rename(columns={'Auto Pilot':'autoPilot'})
    
    # Keep
    dbmn = dbmn[['couponRate','referralRate','buyBoxRepricePercent',
                 'minMarkup','maxMarkup','discountRateforGiftCard',
                 'taxPercent','otherCosts','Vendor','Marketplace',
                 'Inventory','subscriptionDiscount','autoPilot','Issues']]
    #%%
    
    
    
    
    #############################################
    ### Import and join Manual Entry Database ###
    #############################################
    # Read
    manualprices = bf.read_gsheet(client=client, worksheet="Listed Catalog", tab='manualPrices', twoheader=False)
    
    # Format
    manualprices['manual'] = pd.to_numeric(manualprices['manual'])
    manualprices['maxOverride'] = pd.to_numeric(manualprices['maxOverride'])
    #%%
    

    
    ###################################
    ### Start with WMUS_GetAllItems ###
    ###################################
    #%%
    #-- Only SKUs from here are listed
    wmus = bf.read_gsheet(client=client, worksheet='WMUS_GetAllItems', tab='wm-us')
    
    # Filter
    wmus = wmus[wmus['lifecycleStatus']=='ACTIVE']
    wmus['Vendor'] = wmus['sku'].str.split('-').str[0]
    wmus = wmus[wmus['Vendor'].isin(['HDUS','LOUS'])]
    wmus['vendorSKU'] = wmus['sku'].str.split('-').str[-1].str.split('x').str[0]
    
    # Rename
    wmus = wmus.rename(columns={'price.amount':'ourCurrentPrice','sku':'SKU'})
    
    # Keep
    wmus = wmus[['SKU','publishedStatus','unpublishedReasons','ourCurrentPrice','vendorSKU','Vendor']]
    
    # Format
    wmus.replace("No Price",np.nan,inplace=True)
    
    
    
    
    
    #################################
    ### Next add WMUS Item Report ###
    #################################
    #%%
    #-- A few fields not included in main file, updated every two weeks
    wmus_item = bf.wm_item_report(drive=drive,folder=folder,country='us')
    
    # Filter
    wmus_item = wmus_item[wmus_item['Lifecycle Status']=='ACTIVE']
    wmus_item['itemID'] = wmus_item['Item ID'].astype(str)
    wmus_item['Vendor'] = wmus_item['SKU'].str.split('-').str[0]
    wmus_item = wmus_item[wmus_item['Vendor'].isin(['HDUS','LOUS'])]
    
    # Format & Rename
    wmus_item['wmUPC'] = wmus_item['UPC'].astype(str).str.zfill(12)
    wmus_item = wmus_item.rename(columns={
        'Reviews Count':'reviewsCount',
        'Offer Start Date':'offerStartDate',
        'Item Page URL':'wmURL',
        'Price':'ourCurrentPrice',
        'Brand':'wmBrand',
        })
    
    # Keep
    wmus_item = wmus_item[[
        'SKU',
        'itemID',
        'wmURL',
        'reviewsCount',
        'wmUPC',
        'offerStartDate',
        'ourCurrentPrice',
        'Vendor',
        'wmBrand',
        ]]
    #%%
    
    
    
    
    
    #######################################
    ### Import My Data Provider G Sheet ###
    #######################################
    #%%
    #-- Additional offers from marketplace
    # Read
    query = ('SELECT * FROM `carmen-8387920.raw_data.mydataprovider_wm_us`')
    
    mdp = bqclient.query(query).to_dataframe().drop_duplicates()
    
    # Rename
    mdp = mdp.rename(columns={'sku':'SKU',
                              'Offer1Price':'offer1Price',
                              'Offer1Edd':'offer1Edd',
                              'Offer1Seller':'offer1Seller'})
    
    # Keep
    mdp = mdp[['SKU','offer1Price','offer1Edd','offer1Seller',
                     'offer2Price','offer2Edd','offer2Seller',
                     'offer3Price','offer3Edd','offer3Seller',
                     'minDiff']]
    
    # Format
    mdp.replace("No Price",np.nan,inplace=True)
    mdp['offer1Price'] = pd.to_numeric(mdp['offer1Price'])
    mdp['offer2Price'] = pd.to_numeric(mdp['offer2Price'])
    mdp['offer3Price'] = pd.to_numeric(mdp['offer3Price'])
    mdp['offer1Edd'] = pd.to_datetime(mdp['offer1Edd']).dt.date
    mdp['offer2Edd'] = pd.to_datetime(mdp['offer2Edd']).dt.date
    
    # Remove dups by taking out null offer1price
    mdp = mdp.sort_values(by='offer1Price').drop_duplicates(subset='SKU',keep='first')
    #%%
    
    
    
    
    ####################################
    ### Walmart API Logs Item Search ###
    ####################################
    #%%
    #-- only using for blank offer price 1 on mdp
    api = bf.read_gsheet(client=client, worksheet='Walmart API Logs', tab='Item_Search')
    
    # Rename
    api = api.rename(columns={'price_amount':'offer1Price','UPC/GTIN':'wmUPC','itemId':'itemID'})
    
    # Keep
    api = api[['offer1Price','wmUPC','itemID']]
    
    # Format
    api['offer1Price'] = pd.to_numeric(api['offer1Price'])
    #%%
    
    
    
    
    ###############################################
    ### Import and join API Logs for units sold ###
    ###############################################
    #%%
    # Read
    apilog_base = bf.read_gsheet(client=client,worksheet="Walmart API Logs",tab='US_Orders')
    # Rename
    apilog_base.rename(columns={'qty':'unitsSold','sku':'SKU'},inplace=True)
    # Keep
    apilog_base = apilog_base[['orderDate','unitsSold','SKU']]
    # New Fields
    apilog_base['recent'] = np.where(((datetime.now() - pd.to_datetime(apilog_base['orderDate'])).dt.days)<=14,1,0)
    # Format
    apilog_base['unitsSold'] = pd.to_numeric(apilog_base['unitsSold'])
    
    # All Units Dataset
    apilog = pd.DataFrame(apilog_base.groupby('SKU')['unitsSold'].sum().reset_index())
    
    # Last 14 Days Dataset
    recent = pd.DataFrame(apilog_base[apilog_base['recent']==1].groupby('SKU')['unitsSold'].sum().reset_index()).rename(columns={'unitsSold':'unitsSold14Days'})
    
    #%%
    
    
    
    
    ######################################################
    ### Read from past listed to retain column Deleted ###
    ######################################################
    #%%
    # Read
    listed = bf.read_gsheet(client=client, worksheet='Listed', tab='wm-us-listed')
    
    # Rename
    listed.rename(columns={'sku':'SKU'},inplace=True)
    #%%
    
    
    


    #-----------------------#
    #-----------------------#
    #---   Vendor Data   ---#
    #-----------------------#
    #-----------------------#
    
    #######################################
    ### Import HD US Daily Scan G Sheet ###
    #######################################
    #%%
    # Read
    hdus_scan = bf.read_gsheet(client=client,worksheet="HD US Daily Scan",tab='DailyScan')
    
    # Rename
    hdus_scan.rename(columns={'sku':'vendorSKU',
                              'online price 10001':'vendorItemCost',
                              'shipcost':'vendorShipCost',
                              'url':'vendorURL',
                              'vendor maincat':'vendorMainCat',
                              'vendor subcat':'vendorSubCat',
                              'MinQty':'vendorMinQty',
                              'excluded ship states':'excludedShipStates',
                              'Daily Scan Checked':'dailyScanChecked',
                              'vendor ship method':'vendorShipMethod',
                              },inplace=True)
    
    # Keep
    hdus_scan = hdus_scan[[
        'vendorSKU',
        'onlineStock',
        'vendorItemCost',
        'vendorShipCost',
        'vendorURL',
        'vendorMainCat',
        'vendorSubCat',
        'vendorMinQty',
        'excludedShipStates',
        'dailyScanChecked',
        'vendorShipMethod',
        ]]
    
    # Format & new var
    hdus_scan['Vendor'] = 'HDUS'
    #%%
    
    
    
    
    
    #******************************************#
    #*** Alert!                             ***#
    #*** Sku List does not match Daily Scan ***#
    #******************************************#
    newskus = bf.read_gsheet(client=client,worksheet="HD US Daily Scan",tab='newSkuList')
    skualert = pd.merge(hdus_scan[['vendorSKU']].rename(columns={'vendorSKU':'sku'}),newskus[['sku']],on='sku',how='outer',indicator=True)
    hdus_only = pd.to_numeric(len(skualert[skualert['_merge']=='left_only']))
    list_only = pd.to_numeric(len(skualert[skualert['_merge']=='right_only']))
    both_sheets = pd.to_numeric(len(skualert[skualert['_merge']=='both']))
    if (hdus_only + list_only)/(hdus_only+list_only+both_sheets) > 0.2:
        bf.slack_skulist(vendor="HDUS",scan=hdus_only,skulist=list_only,both=both_sheets)
    
    #*********************************************#
    #*** Alert!                                ***#
    #*** Daily Scan has not run in a few hours ***#
    #*********************************************#
    # Inventory on/off due to hd us scan not happening
    current_datetime = pd.Timestamp.now(tz=pytz.timezone('US/Central')).replace(tzinfo=None) # current time where script running
    hdus_scan_time = pd.Timestamp(pd.to_datetime(hdus_scan['dailyScanChecked']).unique()[0]) # last scan time
    warning = round(((current_datetime - hdus_scan_time).total_seconds() / 60),0) # hours between current time and last scan time
    autopilot = dbmn[(dbmn['Marketplace']=='WMUS')&(dbmn['Vendor']=='HDUS')]['autoPilot'].unique()[0].upper().strip() # are we on autopilot?
    curr_inv_stat = dbmn[(dbmn['Marketplace']=='WMUS')&(dbmn['Vendor']=='HDUS')]['Inventory'].unique()[0].upper().strip() # what is the current inventory status?
    # Send alert when conditions are right
    if warning > 720 and autopilot == 'ON' and curr_inv_stat == 'ON':
        warning = int(warning/60)
        bf.slack_dailyscans(script='HDUS Daily Scan',hour=warning,status='Off',marketplace='WMUS')
        dbmn['Inventory'] = np.where((dbmn['Marketplace']=='WMUS')&(dbmn['Vendor']=='HDUS'),'Off',dbmn['Inventory'])
        sheet = client.open('Listed Catalog').worksheet(('values')) # Open up sheet for listing
        sheet.update_cell(4,13,"Off") # write all of df to sheet
    if warning < 720 and autopilot == 'ON' and curr_inv_stat == 'OFF':
        warning = int(warning/60)
        bf.slack_dailyscans(script='HDUS Daily Scan',hour=warning,status='On',marketplace='WMUS')
        dbmn['Inventory'] = np.where((dbmn['Marketplace']=='WMUS')&(dbmn['Vendor']=='HDUS'),'On',dbmn['Inventory'])
        sheet = client.open('Listed Catalog').worksheet(('values')) # Open up sheet for listing
        sheet.update_cell(4,13,"On") # write all of df to sheet
    
    
    
    
    
    ##########################################
    ### Import Lowes US Daily Scan G Sheet ###
    ##########################################
    #%%
    # Read
    lous_scan = bf.read_gsheet(client=client, worksheet='Lowes US Data', tab='dailyScan')
    
    # Rename
    lous_scan.rename(columns={'vendorSku':'vendorSKU',
                               'vendorUrl':'vendorURL',
                               'vendorStock':'onlineStock',},
                      inplace=True)
    
    # Keep
    lous_scan = lous_scan[[
        # 'lastVendorScan',
         # 'productId',
         'vendorSKU',
         'vendorURL',
         # 'vendorUPC',
         'vendorItemCost',
         'vendorShipCost',
         'onlineStock',
         'vendorShipMethod',
         'vendorItemTags',
         # 'vendorPromo',
         'vendorMinQty',
         'excludedShipStates',
         'lastVendorScan',
        ]]    
    lous_scan['Vendor'] = 'LOUS'
    #%%
    
    
    
    #**************#
    #*** Alert! ***#
    #**************#
    #%%
    # listed_count = listed[listed['SKU'].str.startswith('LOUS')].shape[0]
    # lowes_count = lous_scan.shape[0]
    # if (abs(lowes_count-listed_count))/listed_count > 0.1:
    #     if dbmn[(dbmn['Vendor']=='LOUS')&(dbmn['Marketplace']=='WMUS')][['Issues']]=="":
    #         bf.slack_listed_lowes()
    #         sheet = client.open('Listed Catalog').worksheet(('values')) # Open up sheet for listing
    #         sheet.update_cell(7,17,"Daily Scan & Listed Discrepancy") # write all of df to sheet
    # else:
    #     sheet = client.open('Listed Catalog').worksheet(('values')) # Open up sheet for listing
    #     sheet.update_cell(7,17,"") # write all of df to sheet
    #%%
    
    
    
    #**************#
    #*** Alert! ***#
    #**************#
    #%%
    loweskus = bf.read_gsheet(client=client, worksheet='Lowes US Data', tab='skuList')
    skualert = pd.merge(lous_scan[['vendorSKU']].rename(columns={'vendorSKU':'sku'}),loweskus[['sku']],on='sku',how='outer',indicator=True)
    lous_only = pd.to_numeric(len(skualert[skualert['_merge']=='left_only']))
    list_only = pd.to_numeric(len(skualert[skualert['_merge']=='right_only']))
    both_sheets = pd.to_numeric(len(skualert[skualert['_merge']=='both']))
    if (lous_only + list_only)/(lous_only+list_only+both_sheets) > 0.2:
        bf.slack_skulist(vendor="LOUS",scan=lous_only,skulist=list_only,both=both_sheets)
    #%%
    
    
    
    #**************#
    #*** Alert! ***#
    #**************#
    #%%
    # Inventory on/off due to hd us scan not happening
    current_datetime = pd.Timestamp.now(tz=pytz.timezone('US/Central')).replace(tzinfo=None)
    lous_scan_time = pd.Timestamp(pd.to_datetime(lous_scan['lastVendorScan']).unique()[0])
    warning = round(((current_datetime - lous_scan_time).total_seconds() / 60),0)
    autopilot = dbmn[(dbmn['Marketplace']=='WMUS')&(dbmn['Vendor']=='LOUS')]['autoPilot'].unique()[0].upper().strip()
    curr_inv_stat = dbmn[(dbmn['Marketplace']=='WMUS')&(dbmn['Vendor']=='LOUS')]['Inventory'].unique()[0].upper().strip()
    if warning > 720 and autopilot == 'ON' and curr_inv_stat == 'ON':
        warning = int(warning/60)
        bf.slack_dailyscans(script='LOUS Daily Scan',hour=warning,status='Off',marketplace='WMUS')
        dbmn['Inventory'] = np.where((dbmn['Marketplace']=='WMUS')&(dbmn['Vendor']=='LOUS'),'Off',dbmn['Inventory'])
        sheet = client.open('Listed Catalog').worksheet(('values')) # Open up sheet for listing
        sheet.update_cell(7,13,"Off") # write all of df to sheet
    if warning < 720 and autopilot == 'ON' and curr_inv_stat == 'OFF':
        warning = int(warning/60)
        bf.slack_dailyscans(script='LOUS Daily Scan',hour=warning,status='On',marketplace='WMUS')
        dbmn['Inventory'] = np.where((dbmn['Marketplace']=='WMUS')&(dbmn['Vendor']=='LOUS'),'On',dbmn['Inventory'])
        sheet = client.open('Listed Catalog').worksheet(('values')) # Open up sheet for listing
        sheet.update_cell(7,13,"On") # write all of df to sheet
    #%%

    
    
    ################################
    ### Join all sheets together ###
    ################################
    #%%
    # Get all items & item report
    df = pd.merge(wmus,wmus_item,on='SKU',how='left')
    df = bf.dup_col(df=df, take='first')
    
    # HDUS Daily Scan
    df = pd.merge(df,hdus_scan,on=['vendorSKU','Vendor'],how='left',indicator=True)
    df.rename(columns={'_merge':'hdus_scan'},inplace=True)
    
    # LOUS Daily Scan
    df = pd.merge(df,lous_scan,on=['vendorSKU','Vendor'],how='left',indicator=True)
    df.rename(columns={'_merge':'lous_scan'},inplace=True)
    df = bf.dup_col(df=df, take='last')
    
    # My Data Provider
    df = pd.merge(df,mdp,on='SKU',how='left')
    
    # Values
    df['Marketplace'] = 'WMUS'
    df = pd.merge(df,dbmn,on=['Vendor','Marketplace'],how='inner')
    
    # Walmart API Logs
    df = pd.merge(df,api,on='itemID',how='left')
    df = bf.dup_col(df=df, take='first')
    
    # Manual Prices
    df = df.merge(manualprices,on='SKU',how='left')
    df = bf.dup_col(df=df, take='first')
    
    # Listed for deleted (Do we need this?)
    df = pd.merge(df,listed[['SKU','Deleted']].fillna(""),on='SKU',how='left')
    
    # API Logs US Orders
    df = pd.merge(df,apilog,on='SKU',how='left') 
    df = pd.merge(df,recent,on='SKU',how='left')
    #%%
    
    
    time.sleep(60)
    

    #-------------------------------------------#
    #-------------------------------------------#
    #---   Rules, Formulas, & Filters Data   ---#
    #-------------------------------------------#
    #-------------------------------------------#
    
    ############################
    ### deleteNotes & Delete ###
    ############################
    #%%
    # Reset Delete & Delete Notes for new run
    df['Notes'] = ""
    df['deleteNotes'] = ""
    df['Delete'] = ""
    
    # Sales unauthroized for item BLACKLIST
    df['deleteNotes'] = np.where(df['unpublishedReasons'].str.upper().str.strip() == 'SALES UNAUTHORIZED FOR ITEM','Sales unauthorized for item',df['deleteNotes'])
    df['Delete'] = np.where(df['unpublishedReasons'].str.upper().str.strip() == 'SALES UNAUTHORIZED FOR ITEM','Delete',df['Delete'])
    
    # Publish status = system problem then status reason BLACKLIST
    df['deleteNotes'] = np.where(df['publishedStatus'].str.upper().str.strip() == 'SYSTEM_PROBLEM',df['publishedStatus'],df['deleteNotes'])
    df['Delete'] = np.where(df['publishedStatus'].str.upper().str.strip() == 'SYSTEM_PROBLEM',"Delete",df['Delete'])
    
    # Flagged items
    df['deleteNotes']  = np.where(df['unpublishedReasons'].str.upper().str.strip()=='YOUR ITEM HAS BEEN FLAGGED BY OUR INTERNAL TEAM. TO FIND OUT WHY, FILE A CASE IN CASE MANAGEMENT.','Item flagged by WM',df['deleteNotes'])
    df['Delete']  = np.where(df['unpublishedReasons'].str.upper().str.strip()=='YOUR ITEM HAS BEEN FLAGGED BY OUR INTERNAL TEAM. TO FIND OUT WHY, FILE A CASE IN CASE MANAGEMENT.','Delete',df['Delete'])
        
    # NotFound/Discontinued ship method
    df['deleteNotes'] = np.where(df['vendorShipMethod'].str.upper().str.strip()=='NOTFOUND/DISCONTINUED','NotFound/Discontinued',df['deleteNotes'])
    df['Delete'] = np.where(df['vendorShipMethod'].str.upper().str.strip()=='NOTFOUND/DISCONTINUED','Delete',df['Delete'])
    
    # Not in HDUS Daily Scan
    df['Notes'] = np.where((df['hdus_scan']=='left_only')&(df['Vendor']=='HDUS'),"Not in HDUS Daily Scan",df['Notes'])
    df.drop(columns=['hdus_scan'],inplace=True)
    
    # Not in LOUS Daily Scan
    df['Notes'] = np.where((df['lous_scan']=='left_only')&(df['Vendor']=='LOUS'),"Not in LOUS Daily Scan",df['Notes'])
    df.drop(columns=['lous_scan'],inplace=True)
    #%%
    
    
    


    #******************************************#
    #*** Alert!                             ***#
    #*** Shipping Not Found or Discontinued ***#
    #******************************************#
    #%%
    # Ship Method - NotFound/Discontinued
    listcat = bf.read_gsheet(client=client, worksheet='Listed Catalog', tab='wm-us')
    listcat = listcat[['SKU','Delete']]
    listcat = listcat.loc[:,~listcat.columns.duplicated()].copy()
    shipalert1 = listcat[listcat['Delete']==""][['SKU']].dropna().drop_duplicates()
    shipalert2 = df[df['deleteNotes'].str.upper().str.strip()=='NOTFOUND/DISCONTINUED'][['SKU']]
    shipalert3 = pd.merge(shipalert1,shipalert2,on='SKU',how='inner')
    if len(shipalert3)>0:
        bf.slack_shipalerts(file_name='wmus-listed', sku_list=shipalert3)
    #%%
    
    
    
    
    #############################
    ### New Columns & Renames ###
    #############################
    #%%
    # Blank Vendor Item Cost - moving to NA because rules are happening on our min being 0
    # df['vendorItemCost'] = df['vendorItemCost'].fillna(0)
    df['Notes'] = np.where((df['Notes']=="")&(df['vendorItemCost'].isna()),"No Vendor Item Cost",df['Notes'])
    
    # Subscriptions
    df['vendorItemTags'] = df['vendorItemTags'].fillna("")
    df['vendorItemCost'] = np.where(df['vendorItemTags'].str.upper().str.strip()=='SUBSCRIPTION',pd.to_numeric(df['vendorItemCost'])*0.95,pd.to_numeric(df['vendorItemCost']))
    
    # Multipacks
    df['Multipacks'] = pd.to_numeric(df['SKU'].str.split('-').str[-1].str.split('x').str[1].fillna(1))
    df['vendorItemCost'] = pd.to_numeric(df['vendorItemCost']) * pd.to_numeric(df['Multipacks'])
    df['daysActive'] = (datetime.now() - pd.to_datetime(df['offerStartDate'])).dt.days
    df['vendorSKU'] = df['vendorSKU'].str.split('x').str[0]
    
    # onlineStockSafety - divided by 2, no less than 5
    df['onlineStock'] = pd.to_numeric(df['onlineStock'])
    df['onlineStockSafety'] = np.where(df['onlineStock']/2<5,0,df['onlineStock']/2)
    df['onlineStockSafety'] = np.where((df['Vendor']=='LOUS')&(df['onlineStock'].between(1,10,inclusive='both')),15,df['onlineStockSafety'])

    df['Notes'] = np.where((df['onlineStock']>0) & (df['Notes']=="") & (df['onlineStock']/2<5),"Low Stock",df['Notes'])
    
    # onlineStockSafety - greater than 500 set to 500
    df['onlineStockSafety'] = np.round(np.where(df['onlineStockSafety']>500,500,df['onlineStockSafety']),0)
    
    # Fill nan with 0
    df['onlineStockSafety'] = df['onlineStockSafety'].fillna(0)
    
    # When vendor item cost is 0, hten online stock safety is 0
    df['onlineStockSafety'] = np.where(df['vendorItemCost']==0,0,df['onlineStockSafety'])
    df['Notes'] = np.where((df['Notes']=="") & (df['vendorItemCost']==0),"$0 Vendor Item Cost",df['Notes'])
    
    # Set online stock safety to 3 when the item is new
    df['onlineStockSafety'] = np.where((df['onlineStockSafety']>0)&(df['daysActive'].between(1,2,inclusive='both')),3,df['onlineStockSafety'])
    df['Notes'] = np.where((df['Notes']=="")&(df['onlineStockSafety']>0)&(df['daysActive'].between(1,2,inclusive='both')),"New;Qty 3",df['Notes'])
    
    # Walmart API Logs US Orders
    df['unitsSold'] = df['unitsSold'].fillna(0).replace("",0)
    df['unitsSold14Days'] = df['unitsSold14Days'].fillna(0).replace("",0)
    df['deleteNotes'] = np.where(((datetime.now() - pd.to_datetime(df['offerStartDate'])).dt.days >=14)&(df['unitsSold14Days']==0),"No sale 14 days",df['deleteNotes'])
    df['Delete'] = np.where(((datetime.now() - pd.to_datetime(df['offerStartDate'])).dt.days >=14)&(df['unitsSold14Days']==0),"Delete",df['Delete'])
    #%%
    
    
    
    
    #####################
    ### The Blacklist ###
    #####################
    #%%
    # Brand On Listing
    df['deleteNotes'] = np.where((df['offer1Seller']!="")&(df['offer1Seller'].notna())&(df['offer2Seller']!="")&(df['offer2Seller'].notna())&(df['wmBrand']!="")&(df['wmBrand'].notna())
                            &((df['offer1Seller'].str.upper().str.strip().isin(df['wmBrand'].str.upper().str.strip()))
                            |(df['wmBrand'].str.upper().str.strip().isin(df['offer1Seller'].str.upper().str.strip()))
                            |(df['offer2Seller'].str.upper().str.strip().isin(df['wmBrand'].str.upper().str.strip()))
                            |(df['wmBrand'].str.upper().str.strip().isin(df['offer2Seller'].str.upper().str.strip())))
                            ,"Brand On Listing",df['deleteNotes'])
    df['Delete'] = np.where((df['offer1Seller']!="")&(df['offer1Seller'].notna())&(df['offer2Seller']!="")&(df['offer2Seller'].notna())&(df['wmBrand']!="")&(df['wmBrand'].notna())
                            &((df['offer1Seller'].str.upper().str.strip().isin(df['wmBrand'].str.upper().str.strip()))
                            |(df['wmBrand'].str.upper().str.strip().isin(df['offer1Seller'].str.upper().str.strip()))
                            |(df['offer2Seller'].str.upper().str.strip().isin(df['wmBrand'].str.upper().str.strip()))
                            |(df['wmBrand'].str.upper().str.strip().isin(df['offer2Seller'].str.upper().str.strip())))
                            ,"Delete",df['Delete'])
    
    # Add brand to blacklist
    blbr = bf.read_gsheet(client=client, worksheet='Blacklist', tab='wm-brands')
    newblbr = df[df['deleteNotes']=='Brand On Listing'][['wmBrand']].rename(columns={'wmBrand':'Brand'})
    newblbr['Date added'] = pd.Timestamp.now().date().strftime('%Y-%m-%d')
    newblbr['Notes'] = 'Brand On Listing For WMUS for Offer1 or Offer2 Sellers'
    blbr = pd.merge(blbr,newblbr,on='Brand',how='outer')
    blbr = bf.dup_col(df=blbr, take='first')
    blbr = blbr.fillna("").drop_duplicates()
    bf.write_gsheet(client=client, df=blbr, sheet='Blacklist', tab='wm-brands',clear=False)
    
    # Remove from whitelist
    whitelist = bf.read_gsheet(client=client, worksheet='Whitelist', tab='wm-us')
    blbr_sku = df[df['deleteNotes'].isin(['SYSTEM_PROBLEM','Sale unauthorized for item','Brand On Listing'])][['SKU']]
    whitelist = pd.merge(whitelist,blbr_sku,on='SKU',how='left',indicator=True).query('_merge=="left_only"').drop(columns=('_merge')).drop_duplicates()
    bf.write_gsheet(client=client, df=whitelist, sheet='Whitelist', tab='wm-us',clear=False)
    
    # Read in blacklist
    blacklist = bf.read_gsheet(client=client, worksheet="Blacklist", tab='wmIds', twoheader=False)
    # Keep item id from blacklist
    bl = blacklist[['Item ID']].dropna().drop_duplicates()
    # Read (Item Search again for new fields)
    wmapilog = bf.read_gsheet(client=client, worksheet='Walmart API Logs', tab='Item_Search')
    # Rename to match other cols
    wmapilog = wmapilog.rename(columns={'itemId':'Item ID','title':'Item name'})
    # Keep these three
    wmapilog = wmapilog[['Item ID','Item name','SKU']].drop_duplicates()
    # New df for delete items
    delete = pd.DataFrame(data=df[df['deleteNotes'].isin(['SYSTEM_PROBLEM','Sale unauthorized for item','Brand On Listing'])][['itemID','deleteNotes','SKU']],columns=['itemID','deleteNotes','SKU'])
    # Rename
    delete = delete.rename(columns={'itemID':'Item ID','deleteNotes':'Retire Reason'})
    # Format
    delete['Item ID'] = delete['Item ID'].astype(str)
    # Merge
    delete = pd.merge(delete,wmapilog,on='SKU',how='left')
    delete = bf.dup_col(df=delete,take='last')
    # Keep
    delete = delete[['Item ID','Retire Reason','Item name']]
    blacklist = blacklist[['Item ID','Retire Reason','Item name']]
    # Concat
    blacklist = pd.concat([blacklist,delete]).drop_duplicates().fillna("")
    # Any new items added?
    delete = delete[~delete['Item ID'].isin([bl['Item ID'].to_list()])].dropna()
    #%%
    
    
    
    #******************************************#
    #*** Alert!                             ***#
    #*** Retire System Problem to blacklist ***#
    #******************************************#
    #%%
    #*** For Testing ***#
    # Create a new row with specific values
    # new_row = {'Item ID': '123', 'Retire Reason': 'test', 'Item name': 'Test item'}
    
    # Append the new row to the DataFrame
    # delete = delete.append(new_row, ignore_index=True)
    if delete.shape[0]>0:
        for i,r in delete.iterrows():
            sku = r['Item ID']
            reason = r['Retire Reason']
            file_name = 'wmus_listed'
            check = bf.read_gsheet(client=client, worksheet='Blacklist', tab='slack_alert')
            if sku not in check['sku'].to_list():
                # Write to list
                add = {'sku':sku,'reason':reason,'catalog':file_name}
                check = check.append(add,ignore_index=True).fillna("")
                bf.write_gsheet(client=client, df=check, sheet='Blacklist', tab='slack_alert',clear=False)
                bf.slack_blacklist_brands(file_name,sku,reason)
            # if reason=='SYSTEM_PROBLEM':
            #     bf.slack_blacklist_brands(file_name, sku, reason)
    #%%
    
    
    
    
    
    ####################
    ### Calculations ###
    ####################
    #%%
    # Set blank price notes
    df['pricingNotes']                      = ""
    
    # Vendor Costs
    df['vendorShipCost']                    = pd.to_numeric(df['vendorShipCost']).fillna(0)
    df['vendorCostBeforeTaxAndCoupons']     = df['vendorShipCost'] + pd.to_numeric(df['vendorItemCost'])
    df['couponAmount']                      = df['vendorCostBeforeTaxAndCoupons'] * df['couponRate']
    df['vendorCostBeforeTax']               = df['vendorCostBeforeTaxAndCoupons'] - df['couponAmount']
    df['vendorTaxAmount']                   = df['vendorCostBeforeTax'] * df['taxPercent']
    df['vendorCostBeforeDiscounts']         = df['vendorCostBeforeTax'] + df['vendorTaxAmount']
    df['discountAmountForGiftCard']         = df['vendorCostBeforeDiscounts'] * df['discountRateforGiftCard']
    df['vendorCostTotal']                   = df['vendorCostBeforeDiscounts'] - df['discountAmountForGiftCard']
    df['totalCostBeforeReferral']           = df['vendorCostTotal'] + df['otherCosts']
    df['minPriceBeforeReferralFee']         = df['totalCostBeforeReferral']*(1+df['minMarkup'])
    df['minReferralFee']                    = df['minPriceBeforeReferralFee']*(df['referralRate'] / (1-df['referralRate']))
    df['maxPriceBeforeReferralFee']         = df['totalCostBeforeReferral']*(1+df['maxMarkup'])
    df['maxReferralFee']                    = df['maxPriceBeforeReferralFee']*(df['referralRate'] / (1-df['referralRate']))
    
    # Min Max
    df['ourMinPrice']                       = df['minReferralFee'] + df['minPriceBeforeReferralFee']
    df['ourMaxPrice']                       = df['maxReferralFee'] + df['maxPriceBeforeReferralFee']
    df['maxOffer2Offer3']                   = np.maximum(df['offer2Price'],df['offer3Price'])
    df['ourMaxPrice']                       = np.where(pd.to_numeric(df['ourMaxPrice'])<pd.to_numeric(df['maxOffer2Offer3']),pd.to_numeric(df['maxOffer2Offer3']),pd.to_numeric(df['ourMaxPrice']))

    # BB Prices
    df['buyBoxTotalPrice']                  = np.where(df['offer1Price'].isna(),df['ourMinPrice'],pd.to_numeric(df['offer1Price'])) # if it's blank in both api & mdp then take min as safety feature
    df['buyBoxWinPrice']                    = np.where(pd.to_numeric(df['ourCurrentPrice'])==pd.to_numeric(df['buyBoxTotalPrice']),pd.to_numeric(df['ourCurrentPrice']),pd.to_numeric(df['buyBoxTotalPrice']) * (1-df['buyBoxRepricePercent']))
    df['pricingNotes']                      = np.where(pd.to_numeric(df['ourCurrentPrice'])==pd.to_numeric(df['buyBoxTotalPrice']),"Our Current Price is the Buy Box Total Price",df['pricingNotes'])
    
    # Delivery & Pickup
    df['buyBoxOwner']                       = np.where(df['offer1Seller'].str.upper().str.strip()=='THEMARKET',"Yes","No")
    try:
        df['deliveryDateDiff']                  = (df['offer1Edd'].replace("",np.nan).fillna(pd.Timestamp("today").date()) - df['offer2Edd'].replace("",np.nan).fillna(pd.Timestamp("today").date())).dt.days
    except:
        df['deliveryDateDiff'] = ""
    df['pickupItemPrice']                   = np.where((df['offer1Seller'].str.upper().str.strip()=='WALMART.COM')& \
                                                      (df['offer1Edd'].isna())& \
                                                      (df['offer2Seller'].str.upper().str.strip()=='THEMARKET')& \
                                                       ((df['offer3Price']-df['offer2Price'])/df['offer2Price']>0.005) \
                                                      ,df['offer3Price']*0.99,df['buyBoxWinPrice'])
        
    df['pickupItemPrice']                   = np.where((df['offer1Seller'].str.upper().str.strip()=='WALMART.COM')& \
                                                      (df['offer1Edd'].isna())& \
                                                      (df['offer2Seller'].str.upper().str.strip()!='THEMARKET') \
                                                      ,df['offer2Price']*0.99,df['pickupItemPrice'])
        
    df['pickupItemPrice']                   = np.where((df['offer1Seller'].str.upper().str.strip()=='WALMART.COM')& \
                                                      (df['offer1Edd'].isna())& \
                                                      (df['offer2Seller']=="") \
                                                      ,df['ourMinPrice'],df['pickupItemPrice'])
        
    df['wmIgnorePrice']                     = np.where((df['offer1Seller'].str.upper().str.strip()=='WALMART.COM')& \
                                                      (df['offer1Edd'].notna())& \
                                                      (df['offer2Seller'].str.upper().str.strip()=='THEMARKET')& \
                                                      ((df['offer3Price']-df['offer2Price'])/df['offer2Price']>0.005) \
                                                      ,df['offer3Price']*0.99,df['pickupItemPrice'])
        
    df['wmIgnorePrice']                   = np.where((df['offer1Seller'].str.upper().str.strip()=='WALMART.COM')& \
                                                      (df['offer1Edd'].notna())& \
                                                      (df['offer2Seller'].str.upper().str.strip()!='THEMARKET') \
                                                      ,df['offer2Price']*0.99,df['wmIgnorePrice'])
        
    df['wmIgnorePrice']                   = np.where((df['offer1Seller'].str.upper().str.strip()=='WALMART.COM')& \
                                                      (df['offer1Edd'].notna())& \
                                                      (df['offer2Seller']=="") \
                                                      ,df['ourMinPrice'],df['wmIgnorePrice'])
        
    df['wmIgnorePrice']                   = np.where((df['offer1Seller'].str.upper().str.strip()=='WALMART.COM')& \
                                                      (df['offer1Edd'].notna())& \
                                                      (df['offer2Seller'].str.upper().str.strip()=='THEMARKET')& \
                                                      (df['offer3Seller']=="") \
                                                      ,df['ourMaxPrice'],df['wmIgnorePrice'])
        
    # Listing Prices
    df['ourMinPrice']                       = pd.to_numeric(df['ourMinPrice'])
    df['offer1Price']                       = pd.to_numeric(df['offer1Price'])
    df['offer1DiffOurMin']                  = df['offer1Price'] - df['ourMinPrice']
    df['offer2DiffOffer1']                  = df['offer2Price'] - df['offer1Price']
    df['offer2PercOffer1']                  = df['offer2DiffOffer1']/df['offer2Price']
    df['offer2CloseGapPrice']               = np.where((df['buyBoxOwner']=='Yes')&(df['offer2PercOffer1']>=0.005),df['offer2Price']*0.99,df['wmIgnorePrice'])
    df['pricingNotes']                      = np.where((df['buyBoxOwner']=='Yes')&(df['offer2PercOffer1']>=0.005),"Closed the gap to offer 2 price",df['pricingNotes'])
    df['estimatedListingPrice']             = np.where(df['offer2CloseGapPrice']<df['ourMinPrice'],df['ourMinPrice'],df['offer2CloseGapPrice']) # safety factor to ensure no drop below min
    df['pricingNotes']                      = np.where(df['offer2CloseGapPrice']<df['ourMinPrice'],"Our Min Price is higher than the calculated price",df['pricingNotes'])
    df['buyBoxWinner']                      = np.where(df['estimatedListingPrice']<df['buyBoxTotalPrice'],"Yes","No")
     
    # Price/Cost Factors (based off of estimatedListingPrice - offer1 & min)
    df['Estimated Referral Fee']            = df['estimatedListingPrice'] * df['referralRate']
    df['Estimated Item Cost']               = df['Estimated Referral Fee'] + df['totalCostBeforeReferral']
    df['Estimated Profit Per Item']         = df['estimatedListingPrice'] - df['Estimated Item Cost']
    df['estimatedProfitMargin']             = df['Estimated Profit Per Item'] / df['estimatedListingPrice']
    df['estimatedMarkup']                   = df['Estimated Profit Per Item'] / df['totalCostBeforeReferral']
    df['Markup Exceeds 100%']               = np.where(df['estimatedMarkup']>1,"Yes","No")
    # df['markupException']                   = np.where(df['unitsSold']>=10)&(df['wmReviews']>=10)&(df['offer2Diff'])
    df['suggestedPrice']                    = np.where(df['Markup Exceeds 100%']=="No",df['estimatedListingPrice'],(df['totalCostBeforeReferral'] + df['totalCostBeforeReferral'] + (df['totalCostBeforeReferral']*30/85))) # safety factor to ensure price does not exceed markup x2
    df['pricingNotes']                      = np.where(df['Markup Exceeds 100%']=="Yes","The markup exceeds 100%",df['pricingNotes'])
    
    # Final Price/Cost Factors (based off of suggestedPrice)
    df['Final Referral Fee']                = df['suggestedPrice'] * df['referralRate']
    df['Final Item Cost']                   = df['Final Referral Fee'] + df['totalCostBeforeReferral']
    df['Final Profit']                      = df['suggestedPrice'] - df['Final Item Cost']
    df['finalProfitMargin']                 = df['Final Profit'] / df['suggestedPrice']
    df['finalMarkup']                       = df['Final Profit'] / df['totalCostBeforeReferral']
    df['lastUpdated']                       = pd.Timestamp.now(tz=pytz.timezone('US/Central')).replace(tzinfo=None).strftime('%Y-%m-%d %I:%M:%p')
    
    # onlineStockSafety at 0 if ourMinPrice is 0 or missing
    df['onlineStockSafety']                 = np.where(df['ourMinPrice'].isnull(),0,df['onlineStockSafety'])
    df['Notes']                             = np.where((df['Notes']=="") & (df['ourMinPrice'].isnull()),"Null Min Price",df['Notes'])
    
    # Lowes only online stock safety to 15 when between 1 and 10
    df['onlineStockSafety']                 = np.where((df['Vendor']=='LOUS')&(df['onlineStockSafety'].between(1,10,inclusive='both')),15,df['onlineStockSafety'])
    
    # onlineStockSafety Manual Circuit Breaker
    df['onlineStockSafety']                 = np.where(df['Inventory']=="Off",0,df['onlineStockSafety'])
    
    # Manual Price Work    
    df['suggestedPrice']                    = np.where(df['manual']>=pd.to_numeric(df['ourMinPrice'].fillna(0).replace("",0)),df['manual'],df['suggestedPrice'])
    df['ourMaxPrice']                       = np.where(df['maxOverride']>=df['ourMinPrice'],df['maxOverride'],df['ourMaxPrice'])
    
    # Price Rulle
    df['priceRule'] = "Standard"
    df['priceRule'] = np.where(df['manual'].isna(),df['priceRule'],"No Strategy/Delete from Repricer")
    
    # Max Price Rule for Losing BB
    df['minDiff'] = pd.to_numeric(df['minDiff'])
    # df['ourMaxPrice'] = np.where((df['onlineStockSafety']>0)
    #                               &(df['minDiff']>2)
    #                               &(df['offer1Edd']!="")
    #                               &(df['offer1Edd']!=np.NaN)
    #                               &(df['offer1Seller']!="TheMarket")
    #                               &((pd.to_datetime(df['offer1Edd'])-pd.Timedelta(days=3))>pd.to_datetime(datetime.today()))
    #                               ,df['offer1Price']*0.87,df['ourMaxPrice']
    #                               )
    df['ourMaxPrice'] = np.where(df['ourMaxPrice']<df['ourMinPrice'],df['ourMinPrice'],df['ourMaxPrice'])
    
    # Manual Price Join
    manualprices = manualprices.merge(df[['SKU','wmURL','itemID','onlineStock','onlineStockSafety','ourMinPrice','ourMaxPrice']],on='SKU',how='left')
    manualprices = bf.dup_col(df=manualprices, take='last')
    manualprices['manual'] = pd.to_numeric(manualprices['manual'])
    manualprices['maxOverride'] = pd.to_numeric(manualprices['maxOverride'])
    manualprices = manualprices[['marketplace',
                                  'SKU',
                                  'manual',
                                  'wmURL',
                                  'date checked', # Change to date checked
                                  'notes', # Change to Notes
                                  'ourMinPrice',
                                  'ourMaxPrice',
                                  'onlineStock',
                                  'onlineStockSafety',
                                  'itemID',
                                  'maxOverride',
                                  'shippingTemplateId',
                                  ]].drop_duplicates().fillna("")
    #%%
    
    #*********************#
    #** Ship Cost Alert **#
    #*********************#
    shipcost = df[df['vendorShipCost']>0][['SKU','wmBrand','offer1Seller','reviewsCount','itemID','vendorShipCost','wmURL']]
    if shipcost.shape[0]>0:
        shipalerts = bf.read_gsheet(client=client, worksheet='Listed Catalog', tab='ShipCosts')
        shipalerts = shipalerts[['SKU']]
        shipcost = pd.merge(shipcost,shipalerts,on='SKU',how='left',indicator=True).query('_merge=="left_only"').drop(columns=('_merge')).drop_duplicates()
        if shipcost.shape[0]>0:
            shipalerts = pd.merge(shipalerts,shipcost,on='SKU',how='outer')
            shipalerts = shipalerts.drop_duplicates().fillna("")
            bf.write_gsheet(client=client, df=shipalerts, sheet='Listed Catalog', tab='ShipCosts',clear=False)
            shipcost = shipcost[~shipcost['SKU'].isna()].drop_duplicates()
            for i,r in shipcost.iterrows():
                sku = r['SKU']
                wmBrand = r['wmBrand']
                Offer1Seller = r['offer1Seller']
                reviewsCount = r['reviewsCount']
                itemID = r['itemID']
                wmURL = r['wmURL']
                sheet = 'wmus_listed'
                bf.slack_ship_alerts(sku=sku,wmBrand=wmBrand,Offer1Seller=Offer1Seller,reviewsCount=reviewsCount,itemID=itemID,wmURL=wmURL,sheet=sheet)



    #***********************#
    #** Ship States Alert **#
    #***********************#
    shipstates = df[['SKU','excludedShipStates','wmURL']]
    shipstates['excludedShipStates'] = shipstates['excludedShipStates'].str.replace(",",";")
    shipstates['shipStateCount'] = np.where(shipstates['excludedShipStates'].str.count(";") > 0, shipstates['excludedShipStates'].str.count(";") + 1, 0)
    shipstates = shipstates[shipstates['shipStateCount']>=25]
    if shipstates.shape[0]>0:
        shipstatessheet = bf.read_gsheet(client=client, worksheet='Listed Catalog', tab='ShipStates')
        shipstates = pd.merge(shipstatessheet,shipstates,on='SKU',how='outer',indicator=True)
        shipstates = bf.dup_col(df=shipstates, take='last')
        shipstates = shipstates.fillna("")
        if shipstates[shipstates["_merge"]=="right_only"].shape[0]>0:
            for i,r in shipstates.iterrows():
                bf.slack_shipstate_counts(sku=r['SKU'],wmurl=r['wmURL'],shipstate_counts=r['shipStateCount'],sheet='WMUS Listed')
        shipstates = shipstates.drop(columns=['_merge']).drop_duplicates().fillna("")
        bf.write_gsheet(client=client, df=shipstates, sheet='Listed Catalog', tab='ShipStates',clear=False)
    
    ################################################################
    ### Formatting monetary values, percentages, and flat values ###
    ################################################################
    #%%
    round2 = [
        'suggestedPrice',
        'ourMinPrice',
        'ourMaxPrice',
        'vendorItemCost',
        'vendorShipCost',
        'vendorCostBeforeTaxAndCoupons',
        'offer1DiffOurMin',
        'couponAmount',
        'vendorCostBeforeTax',
        'vendorTaxAmount',
        'vendorCostBeforeDiscounts',
        'vendorCostTotal',
        'otherCosts',
        'totalCostBeforeReferral',
        'minPriceBeforeReferralFee',
        'maxPriceBeforeReferralFee',
        'minReferralFee',
        'maxReferralFee',
        'vendorItemCost',
        'buyBoxTotalPrice',
        'ourCurrentPrice',
        'buyBoxWinPrice',
        'estimatedListingPrice',
        'Estimated Referral Fee',
        'Estimated Item Cost',
        'Estimated Profit Per Item',
        'Final Referral Fee',
        'Final Item Cost',
        'Final Profit',
        'offer2DiffOffer1',
        'offer2CloseGapPrice',
        'offer1Price',
        ]
    
    round0 = [
        'onlineStockSafety',
        'reviewsCount',
        'onlineStock',
        'unitsSold',
        'deliveryDateDiff',
        ]
    
    roundpct = [
        'couponRate',
        'taxPercent',
        'discountRateforGiftCard',
        'minMarkup',
        'maxMarkup',
        'buyBoxRepricePercent',
        'estimatedProfitMargin',
        'estimatedMarkup',
        'finalProfitMargin',
        'finalMarkup',
        'offer2PercOffer1',
        ]
    
    bf.fix_format(df=df, round2=round2, round0=round0, roundpct=roundpct)
    
    # Fix Dates
    df['offer1Edd']                         = pd.to_datetime(df['offer1Edd']).dt.date.astype(str).replace('NaT',"")
    df['offer2Edd']                         = pd.to_datetime(df['offer2Edd']).dt.date.astype(str).replace('NaT',"")
    #%%
    
    
    
    ########################################
    ### Write to Listed & Listed Catalog ###
    ########################################
    #%%
    # Write to Listed Sheet
    # temp rename for listed
    df1 = df.rename(columns={
        'vendorItemCost':'Vendor Item Cost',
        'vendorShipCost':'Vendor Ship Cost',
        'itemID':'item_id',
        'suggestedPrice':'Suggested Price',
        })
    
    df1 = df1[['SKU',
        'onlineStockSafety',
        'Suggested Price',
        'Vendor Item Cost',
        'Vendor Ship Cost',
        'item_id',
        'Delete',
        'Deleted',
        'lastUpdated',
        'wmUPC',
        'excludedShipStates',
        'ourMinPrice',
        'unitsSold',
        'manual',
        'ourMaxPrice',
        'priceRule',
        'shippingTemplateId',
        ]].fillna("").drop_duplicates()
    
    # Pricing Strategy
    df1['ourMinPrice'] = np.where(df1['priceRule']=='No Strategy/Delete from Repricer',"",df1['ourMinPrice'])
    df1['ourMaxPrice'] = np.where(df1['priceRule']=='No Strategy/Delete from Repricer',"",df1['ourMaxPrice'])
        
    # Write to Listed Catalog
    df2 = df[[
       'lastUpdated',
        'SKU',
        'onlineStockSafety',
        'onlineStock',
        'vendorItemCost',
        'vendorShipCost',
        'suggestedPrice',
        'ourMinPrice',
        'ourMaxPrice',
        'offer1DiffOurMin',
        'offer1Price',
        'offer1Edd',
        'offer1Seller',
        'offer2Price',
        'offer2Seller',
        'offer3Price',
        'wmUPC',
        'daysActive',
        'unitsSold',
        'Notes',
        'vendorURL',
        'wmURL',
        'itemID',
        'vendorShipMethod',
        'Delete',
        'deleteNotes',
        'pricingNotes',
        'unpublishedReasons',
        'excludedShipStates',
        'vendorItemTags',
        'wmBrand',
        'reviewsCount',
        'discountRateforGiftCard',
        'priceRule',
        ]].fillna("").drop_duplicates()
        #%%
    
    
    
    ###########################################
    ### Write to Listed wm-us-listed GSheet ###
    ###########################################
    #%%
    # Formatting Request
    df2['unitsSold'] = df2['unitsSold'].replace(0,"")
    df1['unitsSold'] = df1['unitsSold'].replace(0,"")
    df2['offer1DiffOurMin'] = df2['offer1DiffOurMin'].replace("","No Value")
 
    # Add from whitelist
    whitelist = bf.read_gsheet(client=client, worksheet='Whitelist', tab='wm-us')
    whitelist = pd.merge(whitelist,df1[['SKU']],on='SKU',how='outer').drop_duplicates()
    bf.write_gsheet(client=client, df=whitelist, sheet='Whitelist', tab='wm-us',clear=False)
    
    # Write to Listed wm-us-listed GSheet
    bf.write_gsheet(client=client, df=df1, sheet='Listed', tab='wm-us-listed',clear=False)
    bf.write_gsheet(client=client, df=df2, sheet='Listed Catalog', tab='wm-us',clear=False)
    bf.write_gsheet(client=client, df=manualprices, sheet='Listed Catalog', tab='manualPrices',clear=False)
    bf.write_gsheet(client=client, df=blacklist, sheet='Blacklist', tab='wmIds',clear=False)
    #%%
    

    
    #######################################
    ### Import HD US Daily Scan G Sheet ###
    #######################################
    #%%
    # HD SKU List queried up top!
    dfns = df[df['Vendor']=='HDUS'][['vendorSKU']].rename(columns={'vendorSKU':'sku'})
    dfns['sku'] = dfns['sku'].str.split('x').str[0].str.split('-').str[0]
    dfns['Marketplace'] = 'WMUS'
    
    # Get everything that matches WMUS
    dfns = pd.merge(dfns,newskus,on=['sku','Marketplace'],how='left').drop_duplicates()
    
    # Concat WMUS with AZUS
    newskus = pd.concat([newskus[newskus['Marketplace']!='WMUS'],dfns])
    bf.write_gsheet(client=client, df=newskus, sheet='HD US Daily Scan', tab='newSkuList',clear=False)     
    #%%
    
    
    
    
    ######################################
    ### Import LOUS SKUs to NewSkuList ###
    ######################################
    #%%
    # Lowes SKU List queried up top!      
    lons = df[df['Vendor']=='LOUS'][['vendorSKU']].rename(columns={'vendorSKU':'sku'})
    lons['sku'] = lons['sku'].str.split('x').str[0].str.split('-').str[0]
    
    # lons = lons[~lons['sku'].isin(loweskus['sku'].to_list())]
    
    loweskus = loweskus[loweskus['excludedShipStates']!=""]
    loweskus = pd.concat([loweskus,lons]).fillna("").drop_duplicates()
    loweskus['sku'] = loweskus['sku'].str.split('x').str[0].str.split('-').str[0]
    
    # loweskus = pd.merge(lons,loweskus,on='sku',how='left',indicator=True)
    
    loweskus['Marketplace'] = np.where((loweskus['Marketplace'].isna())|(loweskus['Marketplace']==""),"WMUS",loweskus['Marketplace'])
    bf.write_gsheet(client=client, df=loweskus, sheet='Lowes US Data', tab='skuList',clear=False)        
    #%%
    
    
    
    
    
    #####################################
    ### Write out to Walmart Repricer ###
    #####################################
    #%%
    # try:
        
    # List out all item reports in g drive
    file_name="RepricerTemplate"
    listed = pd.DataFrame(drive.ListFile({'q': "title contains '"+file_name+"'"}).GetList())
    # Delete the file if it already exists
    if os.path.exists(file_name + ".xlsx"):
        os.remove(file_name + ".xlsx")
    # Find US item report & download 
    latest_report_id = listed[listed['title'].str.startswith(file_name)].sort_values(by=['createdDate'],ascending=False)['id'].reset_index().drop(columns='index').iloc[0][0]
    # Create object for this particular file id
    file = drive.CreateFile({'id': latest_report_id}) 
    # Download file using this command and name whatever is in the blank
    file.GetContentFile(file_name+".xlsx")
    # Dataframe to append to excel workbook - comes from
    df_repricer = df1[~((df1['priceRule']=='Standard')&((df1['ourMinPrice']=="")|(df1['ourMaxPrice']=="")))][['SKU','priceRule','ourMinPrice','ourMaxPrice']].drop_duplicates().fillna("")

    
    #**************
    
    # Load the existing Excel file with openpyxl
    book = load_workbook('RepricerTemplate.xlsx')
    
    # Select the sheet to append to
    ws = book['Repricer Bulk Upload']
    
    # Define the start row and column for appending the data
    start_row = 7
    start_col = 4
    
    # Iterate over the DataFrame rows and columns and append to the sheet
    for r, row in enumerate(df_repricer.values):
        for c, value in enumerate(row):
            cell = ws.cell(row=start_row+r, column=start_col+c)
            cell.value = value
    
    # Save the changes to the Excel file
    book.save('RepricerTemplate.xlsx')
    
    # Set the folder ID of the Google Drive folder to upload to
    folder_id = '1frM7s-BL6GFlhX3Hvo4DQmQ1booNuqoM'
    
    # Set the path to the local file to upload
    file_path = 'RepricerTemplate.xlsx'
    
    # Set up Google Drive API credentials and service
    drive_service = build('drive', 'v3', credentials=credentials)
    
    # Search for the file by name
    file_name = os.path.basename(file_path)
    query = f"name='{file_name}' and trashed=false"
    results = drive_service.files().list(q=query, fields='files(id)').execute().get('files', [])
    if not results:
        print(f'File {file_name} not found in folder ID {folder_id}')
        exit()
    file_id = results[0]['id']
    
    # Create a MediaFileUpload object for the local file
    file_metadata = {'name': file_name}
    media = MediaFileUpload(file_path, resumable=True)
    
    # Upload the file to Google Drive
    try:
        file = drive_service.files().update(
            fileId=file_id,
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()
        print(f'File ID: {file.get("id")} updated successfully in folder ID: {folder_id}')
    except HttpError as error:
        print(f'An error occurred: {error}')
        file = None
    
    #*******************
    #%%
    
    
    
    ########################
    ### Send Slack Alert ###
    ########################
    #%%
    bf.slack_message(file_name = 'wmus_listed')
    #%%
    
    
    
if __name__ == "__main__":
    wmus_listed()